import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OLD_ROUTES_PATH = r"D:\Dev\fabric-server-api\fabric-api\routes\wildfly_routes.js"
NEW_ROUTES_PATH = r"D:\Dev\fabric-server-api\fabric-api\routes\wildfly_routes_v2.js"
OUTPUT_XLSX = "wildfly_route_migration_final_qwen.xlsx"

def extract_route_array(js_content):
    """Extract the array from module.exports = [...]"""
    match = re.search(r"module\.exports\s*=\s*(\[.*\]);?", js_content, re.DOTALL)
    return match.group(1) if match else "[]"

def split_routes(array_str):
    """Split the array into individual route objects"""
    routes = []
    depth = 0
    start = None
    in_string = False
    escape = False
    string_char = None

    for i, ch in enumerate(array_str):
        if escape:
            escape = False
            continue
            
        if ch == '\\' and in_string:
            escape = True
            continue
            
        if not in_string:
            if ch in ['"', "'"]:
                in_string = True
                string_char = ch
            elif ch == '{':
                if depth == 0:
                    start = i
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0 and start is not None:
                    routes.append(array_str[start:i+1])
                    start = None
        else:
            if ch == string_char:
                in_string = False
                string_char = None
                
    return routes

def parse_old_route(route_str):
    """Parse old Hapi v18 format route"""
    method_match = re.search(r'method\s*:\s*["\'](\w+)["\']', route_str)
    path_match = re.search(r'path\s*:\s*["\']([^"\']+)["\']', route_str)
    handler_match = re.search(r'handler\s*:\s*\w+\.(\w+)', route_str)
    
    if not (method_match and path_match and handler_match):
        return None
        
    return {
        "method": method_match.group(1).upper(),  # Normalize to uppercase
        "path": path_match.group(1),
        "handler": handler_match.group(1)
    }

def parse_new_route(route_str):
    """Parse new Hapi v19+ format route"""
    method_match = re.search(r'method\s*:\s*["\'](\w+)["\']', route_str)
    path_match = re.search(r'path\s*:\s*["\']([^"\']+)["\']', route_str)
    handler_match = re.search(r'handler\s*:\s*\w+\.(\w+)', route_str)
    
    if not (method_match and path_match and handler_match):
        return None
    
    # Extract description from options
    description = ""
    desc_match = re.search(r'description\s*:\s*["\']([^"\']*)["\']', route_str)
    if desc_match:
        description = desc_match.group(1)
    
    # Extract the entire options block
    options_match = re.search(r'options\s*:\s*\{', route_str)
    if not options_match:
        return {
            "method": method_match.group(1).upper(),  # Normalize to uppercase
            "path": path_match.group(1),
            "handler": handler_match.group(1),
            "description": description,
            "validate_sections": {}
        }
    
    # Find validate block within options
    validate_sections = {}
    
    # Look for validate: { ... }
    validate_match = re.search(r'validate\s*:\s*\{', route_str)
    if validate_match:
        # Extract each validation section with better pattern
        for section in ['params', 'query', 'payload']:
            # More comprehensive pattern to capture the entire section
            section_pattern = rf'{section}\s*:\s*(Joi\.object\([^)]*\{{.*?\}}\s*\)|{{[^{{]*?}})'
            section_match = re.search(section_pattern, route_str, re.DOTALL)
            if section_match:
                validate_sections[section] = section_match.group(1)
    
    return {
        "method": method_match.group(1).upper(),  # Normalize to uppercase
        "path": path_match.group(1),
        "handler": handler_match.group(1),
        "description": description,
        "validate_sections": validate_sections
    }

def extract_nested_joi_object(joi_str, parent_name=""):
    """Recursively extract nested Joi objects with full hierarchy and clean validation"""
    fields = []
    
    # Handle Joi.object({ ... }) wrapper
    inner_content = joi_str
    if 'Joi.object(' in joi_str:
        # Extract the inner object content and the full Joi chain after it
        obj_match = re.search(r'(Joi\.object\s*\(\s*\{.*?\}\s*\))(.*)', joi_str, re.DOTALL)
        if obj_match:
            object_part = obj_match.group(1)
            trailing_chain = obj_match.group(2).strip()
            
            # Extract description from trailing chain
            description = ""
            desc_match = re.search(r'\.description\s*\(\s*["\']([^"\']*)["\']\s*\)', trailing_chain)
            if desc_match:
                description = desc_match.group(1)
                # Remove description from validation
                trailing_chain = re.sub(r'\.description\s*\([^)]*\)', '', trailing_chain)
            
            validation_rules = trailing_chain.strip()
            
            # Add the parent object field
            full_name = parent_name if parent_name else "root"
            if full_name != "root":
                fields.append({
                    "name": full_name,
                    "description": description,
                    "validation": validation_rules if validation_rules else "(object)"
                })
            
            # Now extract inner fields from the object content
            inner_match = re.search(r'Joi\.object\s*\(\s*\{(.*)\}\s*\)', object_part, re.DOTALL)
            if inner_match:
                inner_fields_str = inner_match.group(1)
                inner_fields = parse_joi_fields_flat(inner_fields_str, full_name if full_name != "root" else "")
                fields.extend(inner_fields)
        else:
            # Fallback: treat as flat
            fields = parse_joi_fields_flat(joi_str, parent_name)
    else:
        # Not a Joi.object — treat as flat
        fields = parse_joi_fields_flat(joi_str, parent_name)
    
    return fields

def parse_joi_fields_flat(fields_str, parent_name=""):
    """Parse a flat set of Joi field definitions (not nested objects)"""
    fields = []
    pos = 0
    content = fields_str.strip()
    
    while pos < len(content):
        # Skip whitespace
        while pos < len(content) and content[pos].isspace():
            pos += 1
        if pos >= len(content):
            break
            
        # Match field name: key:
        field_match = re.match(r'(\w+)\s*:\s*', content[pos:])
        if not field_match:
            # Skip to next line or }
            next_brace = content.find('}', pos)
            next_comma = content.find(',', pos)
            if next_brace != -1 and (next_comma == -1 or next_brace < next_comma):
                break
            pos = next_comma + 1 if next_comma != -1 else len(content)
            continue
            
        field_name = field_match.group(1)
        pos += field_match.end()
        
        # Determine field end (find next , or } at top level)
        depth = 0
        in_string = False
        str_char = None
        escape = False
        start = pos
        i = pos
        
        while i < len(content):
            ch = content[i]
            
            if escape:
                escape = False
            elif ch == '\\':
                escape = True
            elif not in_string:
                if ch in ['"', "'"]:
                    in_string = True
                    str_char = ch
                elif ch in ['(', '{', '[']:
                    depth += 1
                elif ch in [')', '}', ']']:
                    depth -= 1
                    if depth < 0:
                        break
                elif ch == ',' and depth == 0:
                    break
            else:
                if ch == str_char:
                    in_string = False
                    str_char = None
            
            i += 1
        
        joi_chain = content[start:i].strip()
        if joi_chain.endswith(','):
            joi_chain = joi_chain[:-1].strip()
        
        # Extract description
        description = ""
        desc_match = re.search(r'\.description\s*\(\s*["\']([^"\']*)["\']\s*\)', joi_chain)
        if desc_match:
            description = desc_match.group(1)
            joi_chain = re.sub(r'\.description\s*\([^)]*\)', '', joi_chain)
        
        # CLEAN VALIDATION: Remove Joi.type() wrapper and keep only validation rules
        validation = joi_chain.strip()
        if validation.startswith('Joi.'):
            # Remove the Joi.type() part (e.g., "Joi.string()" or "Joi.object({...})")
            # Match: Joi.<type>(...)
            type_match = re.match(r'Joi\.\w+\s*\([^)]*\)', validation)
            if type_match:
                validation = validation[type_match.end():].strip()
                # Clean up remaining parentheses/braces from complex types
                validation = re.sub(r'^\s*\)\s*', '', validation)
                validation = re.sub(r'^\s*\}\s*', '', validation)
        
        # If validation is empty after cleanup, mark as none
        if not validation:
            validation = "(none)"
        
        # Build full name
        full_name = f"{parent_name}.{field_name}" if parent_name else field_name
        
        fields.append({
            "name": full_name,
            "description": description,
            "validation": validation
        })
        
        pos = i + 1
    
    return fields

def extract_joi_fields(validation_str, is_payload=False):
    """Extract individual Joi field definitions including nested objects"""
    if not validation_str:
        return []
    
    return extract_nested_joi_object(validation_str)

def main():
    print("=" * 60)
    print("🔍 HAPI ROUTE MIGRATION ANALYZER")
    print("=" * 60)
    
    # Read files
    print(f"\n📖 Reading old routes from: {OLD_ROUTES_PATH}")
    with open(OLD_ROUTES_PATH, "r", encoding="utf-8") as f:
        old_content = f.read()
    
    print(f"📖 Reading new routes from: {NEW_ROUTES_PATH}")
    with open(NEW_ROUTES_PATH, "r", encoding="utf-8") as f:
        new_content = f.read()
    
    # Extract route arrays
    old_array = extract_route_array(old_content)
    new_array = extract_route_array(new_content)
    
    # Split into individual routes
    print("\n🔨 Parsing route objects...")
    old_route_strings = split_routes(old_array)
    new_route_strings = split_routes(new_array)
    
    print(f"   Found {len(old_route_strings)} old route objects")
    print(f"   Found {len(new_route_strings)} new route objects")
    
    # Parse routes
    old_routes = []
    for route_str in old_route_strings:
        parsed = parse_old_route(route_str)
        if parsed:
            old_routes.append(parsed)
    
    new_routes = []
    for route_str in new_route_strings:
        parsed = parse_new_route(route_str)
        if parsed:
            new_routes.append(parsed)
    
    print(f"   Successfully parsed {len(old_routes)} old routes")
    print(f"   Successfully parsed {len(new_routes)} new routes")
    
    # Index new routes by (method, handler) - case insensitive
    new_map = {}
    for route in new_routes:
        key = (route["method"].upper(), route["handler"])
        new_map[key] = route
        print(f"   Indexed: {key} -> {route['path']}")
    
    # Match and extract parameters
    print("\n🔗 Matching routes and extracting parameters...")
    rows = []
    matched_count = 0
    unmatched = []
    
    for old_route in old_routes:
        key = (old_route["method"].upper(), old_route["handler"])
        new_route = new_map.get(key)
        
        if not new_route:
            unmatched.append(f"{old_route['method']} {old_route['path']} -> {old_route['handler']}")
            continue
        
        matched_count += 1
        old_route_str = f"{old_route['method']} {old_route['path']}"
        new_route_str = f"{new_route['method']} {new_route['path']}"
        route_desc = new_route["description"]
        
        print(f"\n   ✓ Matched: {old_route_str} -> {new_route_str}")
        
        # Extract parameters from each validation section
        has_params = False
        for section_name, section_content in new_route["validate_sections"].items():
            is_payload = (section_name == "payload")
            fields = extract_joi_fields(section_content, is_payload)
            
            for field in fields:
                rows.append([
                    old_route_str,
                    new_route_str,
                    route_desc,
                    section_name,
                    field["name"],
                    field["description"],
                    field["validation"]
                ])
                has_params = True
                print(f"      • {section_name}.{field['name']}: {field['validation']}")
        
        if not has_params:
            # Add a single row even if no parameters
            rows.append([
                old_route_str,
                new_route_str,
                route_desc,
                "",
                "",
                "",
                ""
            ])
            print(f"      ⚠️ No parameters found")
    
    # Report results
    print("\n" + "=" * 60)
    print("📊 RESULTS")
    print("=" * 60)
    print(f"✅ Matched routes: {matched_count}")
    print(f"❌ Unmatched routes: {len(unmatched)}")
    print(f"📝 Total parameter rows: {len(rows)}")
    
    if unmatched:
        print("\n⚠️  Unmatched routes:")
        for route in unmatched[:10]:
            print(f"   • {route}")
        if len(unmatched) > 10:
            print(f"   ... and {len(unmatched) - 10} more")
    
    if not rows:
        print("\n❌ ERROR: No parameters extracted!")
        return
    
    # Create Excel file with merged cells for route info
    print(f"\n💾 Creating Excel file: {OUTPUT_XLSX}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Migration"
    
    # Headers with styling
    headers = [
        "Old Route",
        "New Route", 
        "Route Description",
        "Parameter Type",
        "Parameter Name",
        "Parameter Description",
        "Validation Rules"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Group rows by route and parameter type, then merge cells
    current_row = 2
    current_route = None
    current_param_type = None
    route_start_row = 2
    param_type_start_row = 2
    
    for row_data in rows:
        route_key = (row_data[0], row_data[1], row_data[2])  # Old route, new route, description
        param_type = row_data[3]  # Parameter Type
        
        # Write the row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            if col_num == 7:  # Validation Rules column
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Check if we changed routes
        if current_route is not None and route_key != current_route:
            # Merge cells for columns 1-3 (Old Route, New Route, Description)
            if current_row - route_start_row > 1:
                for col_num in range(1, 4):
                    ws.merge_cells(start_row=route_start_row, start_column=col_num,
                                 end_row=current_row - 1, end_column=col_num)
                    ws.cell(row=route_start_row, column=col_num).alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
            
            # Merge parameter type for the last group of previous route
            if current_row - param_type_start_row > 1:
                ws.merge_cells(start_row=param_type_start_row, start_column=4,
                             end_row=current_row - 1, end_column=4)
                ws.cell(row=param_type_start_row, column=4).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            
            route_start_row = current_row
            param_type_start_row = current_row
            current_param_type = param_type
        
        # Check if we changed parameter type within the same route
        elif current_param_type is not None and param_type != current_param_type:
            # Merge parameter type for the previous group
            if current_row - param_type_start_row > 1:
                ws.merge_cells(start_row=param_type_start_row, start_column=4,
                             end_row=current_row - 1, end_column=4)
                ws.cell(row=param_type_start_row, column=4).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            param_type_start_row = current_row
            current_param_type = param_type
        
        current_route = route_key
        if current_param_type is None:
            current_param_type = param_type
        
        current_row += 1
    
    # Merge the last route group (columns 1-3)
    if current_row - route_start_row > 1:
        for col_num in range(1, 4):
            ws.merge_cells(start_row=route_start_row, start_column=col_num,
                         end_row=current_row - 1, end_column=col_num)
            ws.cell(row=route_start_row, column=col_num).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
    
    # Merge the last parameter type group (column 4)
    if current_row - param_type_start_row > 1:
        ws.merge_cells(start_row=param_type_start_row, start_column=4,
                     end_row=current_row - 1, end_column=4)
        ws.cell(row=param_type_start_row, column=4).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    ws.column_dimensions[get_column_letter(7)].width = 80
    
    wb.save(OUTPUT_XLSX)
    print(f"✅ Success! Saved to: {os.path.abspath(OUTPUT_XLSX)}")
    print("=" * 60)

if __name__ == "__main__":
    main()